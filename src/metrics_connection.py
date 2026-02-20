import pandas as pd
import argparse
from sklearn.preprocessing import MinMaxScaler
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter



def connect_mean_value(type_cluster, type_scaffold, generators_name_list, receptor, subset='', data_folder=''):
    """
    Combines mean values from multiple generators into a single DataFrame.
    
    Parameters:
    - receptor: Target receptor
    - type_scaffold: Type of scaffold
    - type_cluster: Cluster type
    - generators_name_list: List of generator names
    - subset: Data subset to be analyzed

    Returns:
    - A combined DataFrame with mean values
    """
    
    # Define path to data
    link = f"{data_folder}data/results/{receptor}/{type_scaffold}_scaffolds/{type_cluster}"
    
    # List to store paths of mean value files
    link_mean = []
    for generator in generators_name_list:
        link_mean.append(f"{link}/{generator}/{generator}_mean_{type_scaffold}_{type_cluster}.csv")

    # Load data and merge into a single DataFrame
    df_list = [pd.read_csv(f) for f in link_mean]
    df = pd.concat(df_list, ignore_index=True)

    # Save results to CSV files
    df.to_csv(f"{data_folder}data/results/{receptor}/{type_scaffold}_scaffolds/{type_cluster}/mean_{type_scaffold}_{type_cluster}{subset}.csv", index=False)

    return df


def connect_mean_value_normalized(type_cluster, type_scaffold, generators_name_list, receptor, subset, data_folder = ''):
    """
    Loads and normalizes mean values using Min-Max scaling.
    
    Parameters:
    - receptor: Target receptor
    - type_scaffold: Type of scaffold
    - type_cluster: Cluster type
    - generators_name_list: List of generator names
    - subset: Data subset to be analyzed

    Returns:
    - Normalized DataFrame
    """

    # Define path to data
    link = f"{data_folder}data/results/{receptor}/{type_scaffold}_scaffolds/{type_cluster}"

    # List to store paths of mean value files
    link_mean = [f"{link}/{generator}/{generator}_mean_{type_scaffold}_{type_cluster}.csv" for generator in generators_name_list]

    # Load data
    df_list = [pd.read_csv(f) for f in link_mean]
    df = pd.concat(df_list, ignore_index=True)

    # Normalize using Min-Max scaling
    scaler = MinMaxScaler()
    numeric_columns = df.select_dtypes(include=['number']).columns  # Select only numeric columns
    df[numeric_columns] = scaler.fit_transform(df[numeric_columns])  # Apply normalization

    # Save normalized results
    df.to_csv(f"{data_folder}data/results/{receptor}/{type_scaffold}_scaffolds/{type_cluster}/mean_{type_scaffold}_{type_cluster}{subset}_norm_min_max.csv", index=False)

    return df


def prettify_generator_name(name: str) -> str:
    if not isinstance(name, str):
        return name

    # remove trailing _mean
    name = name.replace("_mean", "")

    # DrugEx_GT_epsilon_0.1 → DrugEx GT\n epsilon 0.1
    if name.startswith("DrugEx_GT_epsilon_"):
        eps = name.split("DrugEx_GT_epsilon_")[-1]
        return f"DrugEx GT\nepsilon {eps}"
    
    if name.startswith("DrugEx_RNN_epsilon_"):
        eps = name.split("DrugEx_RNN_epsilon_")[-1]
        return f"DrugEx RNN\nepsilon {eps}"

    # GB_GA_mut_r_0.01 → GB_GA\nmut.r. 0.01
    if name.startswith("GB_GA_mut_r_"):
        rate = name.split("GB_GA_mut_r_")[-1]
        return f"GB_GA\nmut.r. {rate}"
    
    if name.startswith("GB_GA_log_p_mut_r_"):
        rate = name.split("GB_GA_log_p_mut_r_")[-1]
        return f"GB_GA log_p\nmut.r. {rate}"

    if name.startswith("addcarbon"):
        return f"AddCarbon"

    # fallback: just remove underscores and keep readable
    return name.replace("_", " ")



def export_article_excel_table(
    df: pd.DataFrame,
    out_xlsx: str,
    receptor: str,
    type_scaffold: str,
    subset: str = "",
    type_cluster: str | None = None,
    ):
    """
    Creates an Excel table suitable for a paper:

    Formatting rules:
      - RS, SED: 2 decimals (Excel format 0.00)
      - ASER: displayed in units of ×10^-2 (i.e., ASER_scaled = ASER * 100),
              3 decimals (Excel format 0.000)
      - Column header for ASER is "ASER · 10^-2"
      - Maximum value in each metric column is bold within each block (Split dis/sim)
      - Header styling, borders, alignment
      - Supports two blocks per sheet (Split = dis/sim). If Split missing, uses type_cluster.
    """

    # --- Safety: ensure folder exists
    out_path = Path(out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # --- Work on a copy
    df = df.copy()

    # Ensure expected columns exist
    if "Scaffold" not in df.columns:
        df["Scaffold"] = type_scaffold

    if "Split" not in df.columns:
        df["Split"] = type_cluster if type_cluster is not None else ""

    if "name" in df.columns and "Name" not in df.columns:
        df = df.rename(columns={"name": "Name"})

    # Prettify generator names
    df["Name"] = df["Name"].apply(prettify_generator_name)
    
    df["Scaffold"] = df["Scaffold"].astype(str).str.upper()


    # Keep and order columns
    wanted_cols = ["Name", "RS", "SED", "ASER", "Scaffold", "Split"]
    existing_cols = [c for c in wanted_cols if c in df.columns]
    df = df[existing_cols]

    metric_cols = [c for c in ["RS", "SED", "ASER"] if c in df.columns]

    # ---- Metric-specific numeric formatting ----
    # RS/SED: 2 decimals
    for c in ["RS", "SED"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)

    # ASER: scale to units of ×10^-2 (multiply by 100) and round to 3 decimals
    if "ASER" in df.columns:
        df["ASER"] = pd.to_numeric(df["ASER"], errors="coerce") * 100.0
        df["ASER"] = df["ASER"].round(3)

    # Header labels for Excel (paper-friendly)
    header_labels = {
        "Name": "Name",
        "RS": "RS",
        "SED": "SED",
        "ASER": "ASER · 10⁻²",   # unit shown in header; values are scaled
        "Scaffold": "Scaffold",
        "Split": "Split",
    }

    # --- Workbook / sheet
    wb = Workbook()
    ws = wb.active
    sheet_name = f"{receptor}_{type_scaffold}"
    ws.title = sheet_name[:31]  # Excel limit

    # Styles
    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Helper: write one block (one split)
    def write_block(start_row: int, block_title: str, block_df: pd.DataFrame) -> int:
        r = start_row
        ncols = len(existing_cols)

        # Title row merged across columns
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1, value=block_title)
        cell.font = title_font
        cell.alignment = left
        r += 1

        # Header row
        for j, col in enumerate(existing_cols, start=1):
            label = header_labels.get(col, col)
            c = ws.cell(row=r, column=j, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        r += 1

        # Maxima per metric within this block
        maxima = {}
        for mc in metric_cols:
            if mc in block_df.columns:
                maxima[mc] = block_df[mc].max(skipna=True)

        # Data rows
        for _, row in block_df.iterrows():
            for j, col in enumerate(existing_cols, start=1):
                val = row[col]
                cell = ws.cell(row=r, column=j, value=val)

                # Alignment
                cell.alignment = center

                # Borders
                cell.border = border

                # Numeric formatting + bold maxima
                if col in metric_cols:

                    cell.number_format = "0.00"

                    # Bold maxima (handle ties)
                    try:
                        if pd.notna(val) and pd.notna(maxima.get(col)) and float(val) == float(maxima[col]):
                            cell.font = bold_font
                    except Exception:
                        pass

            r += 1

        # Blank line after block
        return r + 1

    # --- Split order: dis then sim then anything else
    split_order = ["dis", "sim"]
    unique_splits = df["Split"].astype(str).unique().tolist()
    splits = [s for s in split_order if s in unique_splits]
    splits += [s for s in unique_splits if s not in splits]

    if type_scaffold == 'csk':
        type_scaffold_str = 'CSK'
    elif type_scaffold == 'murcko':
        type_scaffold_str = 'MURCKO'
    else:
        type_scaffold_str = type_scaffold

    # Caption
    current_row = 1
    caption = f"Receptor: {receptor} | Scaffold: {type_scaffold_str}{(' | ' + subset) if subset else ''}"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(existing_cols))
    cap_cell = ws.cell(row=current_row, column=1, value=caption)
    cap_cell.font = Font(bold=True, size=13)
    cap_cell.alignment = left
    current_row += 2

    # Write blocks
    for sp in splits:
        block_df = df[df["Split"].astype(str) == str(sp)].copy()
        block_title = f"Split: {sp}" if sp else "Results"
        current_row = write_block(current_row, block_title, block_df)

    # Column widths
    col_widths = {
        "Name": 22,
        "RS": 10,
        "SED": 10,
        "ASER": 12,      # a bit wider because header includes ·10^-2
        "Scaffold": 10,
        "Split": 8,
    }
    for j, col in enumerate(existing_cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = col_widths.get(col, 12)

    # Freeze panes below caption + header area
    ws.freeze_panes = "A4"

    wb.save(out_xlsx)
    print(f"[OK] Excel saved to: {out_xlsx}")




def main():
    parser = argparse.ArgumentParser(description='Compute and visualize recall metric.')
     # Required arguments
    parser.add_argument('--type_cluster', type=str, required=True, help='Type of clustering (dis/sim)')
    parser.add_argument('--type_scaffold', type=str, required=True, help='Type of scaffold')
    parser.add_argument('--generator_list', nargs='+', required=True, help='Generator name')
    parser.add_argument('--receptor', type=str, required=True, help='Receptor name')
    parser.add_argument('--subset', type=str, required=False, help='Subset')
    parser.add_argument('--data_folder', type=str, required=False, help='Data dir')
    
    args = parser.parse_args()
    
    df_raw = connect_mean_value(args.type_cluster, args.type_scaffold, args.generator_list, args.receptor, args.subset, args.data_folder)
    connect_mean_value_normalized(args.type_cluster, args.type_scaffold, args.generator_list, args.receptor, args.subset, args.data_folder)

    out_xlsx = f"{args.data_folder}data/results/{args.receptor}/{args.type_scaffold}_scaffolds/{args.type_cluster}/paper_table_{args.receptor}_{args.type_scaffold}{args.subset}.xlsx"

    export_article_excel_table(
        df=df_raw,
        out_xlsx=out_xlsx,
        receptor=args.receptor,
        type_scaffold=args.type_scaffold,
        subset=args.subset or "",
        type_cluster=args.type_cluster,
    )




if __name__ == "__main__":
    main()